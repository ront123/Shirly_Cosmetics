#!/usr/bin/env python3
"""
Easybizy → Supabase Import Script
Reads Customers.xlsx and inserts all clients into the clients table.
Run from the project root: python3 backend/scripts/import_customers.py
"""

import openpyxl
import psycopg2
from datetime import datetime

# ── Database connection ─────────────────────────────────────────────────────
DB_URL = "postgresql://postgres:Shirlysony3!!@db.qfzskmsshtxzpahigtab.supabase.co:5432/postgres"

# ── Excel file ──────────────────────────────────────────────────────────────
EXCEL_FILE = "Customers.xlsx"

def parse_name(full_name):
    """Split 'First Last' into (first, last)"""
    if not full_name:
        return ('לקוח', 'לא ידוע')
    parts = str(full_name).strip().split(' ', 1)
    return (parts[0], parts[1]) if len(parts) > 1 else (parts[0], '')

def safe_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    return None

def safe_str(val, max_len=None):
    if val is None:
        return None
    s = str(val).strip()
    if max_len:
        s = s[:max_len]
    return s if s else None

def main():
    print("📂 Reading Customers.xlsx...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    col = {h: i for i, h in enumerate(headers)}
    
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    print(f"✅ Found {len(rows)} customers")
    
    print("🔌 Connecting to Supabase...")
    try:
        conn = psycopg2.connect(DB_URL, connect_timeout=10)
        cursor = conn.cursor()
        print("✅ Connected!")
    except Exception as e:
        print(f"❌ DB Connection failed: {e}")
        print("\n⚠️  Cannot reach Supabase directly from this machine.")
        print("👉 Use the SQL export method instead:")
        generate_sql(rows, col)
        return
    
    inserted = 0
    skipped  = 0
    errors   = 0
    
    for row in rows:
        try:
            full_name = row[col['CustomerName']]
            phone     = safe_str(row[col['MobileFirst']], 20)
            email     = safe_str(row[col['EmailAddress']], 100)
            last_visit = row[col['LastMeeting']]
            total_spent = row[col['TotalSpent']]
            visits    = row[col['HistoryMeetingsCount']]
            gender    = safe_str(row[col['Gender']])
            birth     = row[col['DateOfBirth']]
            created   = row[col['CreatedOn']]
            
            if not phone:
                skipped += 1
                continue
            
            first_name, last_name = parse_name(full_name)
            
            notes_parts = []
            if gender:     notes_parts.append(f"מגדר: {gender}")
            if total_spent: notes_parts.append(f"סה\"כ הוצאה: ₪{total_spent}")
            if visits:     notes_parts.append(f"מספר ביקורים: {visits}")
            if birth:      notes_parts.append(f"תאריך לידה: {safe_date(birth)}")
            notes = ' | '.join(notes_parts) if notes_parts else None
            
            cursor.execute("""
                INSERT INTO clients (first_name, last_name, phone_number, email, notes, last_visit_date, created_at)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (phone_number) DO UPDATE SET
                    first_name = EXCLUDED.first_name,
                    last_name  = EXCLUDED.last_name,
                    email      = COALESCE(EXCLUDED.email, clients.email),
                    notes      = EXCLUDED.notes,
                    last_visit_date = EXCLUDED.last_visit_date
            """, (
                first_name[:100],
                last_name[:100],
                phone,
                email,
                notes,
                last_visit,
                created or datetime.now()
            ))
            inserted += 1
            
            if inserted % 100 == 0:
                conn.commit()
                print(f"  ↳ Imported {inserted} clients...")
                
        except Exception as e:
            errors += 1
            if errors <= 5:
                print(f"  ⚠️  Row error: {e}")
    
    conn.commit()
    cursor.close()
    conn.close()
    
    print(f"\n🎉 Done!")
    print(f"  ✅ Inserted/Updated: {inserted}")
    print(f"  ⏭️  Skipped (no phone): {skipped}")
    print(f"  ❌ Errors: {errors}")


def generate_sql(rows, col):
    """Fallback: generate SQL file to paste into Supabase SQL Editor"""
    print("\n📝 Generating SQL insert file...")
    
    sql_lines = [
        "-- Customers import from Easybizy",
        "-- Paste this in Supabase SQL Editor → Run",
        ""
    ]
    
    count = 0
    for row in rows:
        full_name  = row[col['CustomerName']]
        phone      = row[col['MobileFirst']]
        email      = row[col['EmailAddress']]
        last_visit = row[col['LastMeeting']]
        total_spent = row[col['TotalSpent']]
        visits     = row[col['HistoryMeetingsCount']]
        gender     = row[col['Gender']]
        birth      = row[col['DateOfBirth']]
        created    = row[col['CreatedOn']]
        
        if not phone:
            continue
        
        first_name, last_name = parse_name(full_name)
        phone_clean = str(phone).strip()[:20].replace("'", "")
        first_clean = first_name[:100].replace("'", "''")
        last_clean  = last_name[:100].replace("'", "''")
        email_val   = f"'{str(email)[:100]}'" if email else 'NULL'
        
        notes_parts = []
        if gender:      notes_parts.append(f"מגדר: {gender}")
        if total_spent: notes_parts.append(f"הוצאה: {total_spent}")
        if visits:      notes_parts.append(f"ביקורים: {visits}")
        notes_val = f"'{' | '.join(notes_parts)}'" if notes_parts else 'NULL'
        
        last_visit_val = f"'{last_visit}'" if isinstance(last_visit, datetime) else 'NULL'
        created_val    = f"'{created}'"    if isinstance(created, datetime)    else 'NOW()'
        
        sql_lines.append(
            f"INSERT INTO clients (first_name, last_name, phone_number, email, notes, last_visit_date, created_at) "
            f"VALUES ('{first_clean}', '{last_clean}', '{phone_clean}', {email_val}, {notes_val}, {last_visit_val}, {created_val}) "
            f"ON CONFLICT (phone_number) DO UPDATE SET first_name=EXCLUDED.first_name, email=COALESCE(EXCLUDED.email,clients.email);"
        )
        count += 1
    
    out_file = "backend/scripts/import_customers.sql"
    with open(out_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(sql_lines))
    
    print(f"✅ Generated {out_file} with {count} INSERT statements")
    print(f"👉 Open Supabase SQL Editor and run that file!")


if __name__ == '__main__':
    main()
